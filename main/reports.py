from django.http import HttpResponse
from .models import Course, Abiturient, RecordCourse, Teacher
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

# Oтчет по курсам
def generate_courses_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Курсы"

    headers = [
        '№',
        'Название курса',
        'Преподаватель',
        'Уровень',
        'Цена (₽)',
        'Дата начала',
        'Дата окончания',
        'Расписание',
        'Записано',
        'Мест всего',
        'Свободно',
        'Статус'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

    courses = Course.objects.select_related('teacher').filter(is_activ=True)

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 12

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заполнение данными
    for idx, course in enumerate(courses, start=1):
        enrolled_count = course.get_enrolled_count()
        max_students = course.max_students or 0
        available_spots = max(0, max_students - enrolled_count)

        teacher_name = f"{course.teacher.last_name} {course.teacher.first_name}" if course.teacher else ""

        status = "Активен" if course.is_activ else "Архив"

        ws.append([
            idx,
            course.name or '',
            teacher_name,
            course.level or '',
            float(course.price) if course.price else 0,
            course.start_date.strftime('%d.%m.%Y') if course.start_date else '',
            course.end_date.strftime('%d.%m.%Y') if course.end_date else '',
            course.raspisanie or '',
            enrolled_count,
            max_students,
            available_spots,
            status,
        ])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_courses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)

    return response

# Oтчет по абитуриентам
def generate_students_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Абитуриенты"

    headers = [
        '№',
        'Фамилия',
        'Имя',
        'Email',
        'Телефон',
        'Класс',
        'Школа',
        'Родитель',
        'Телефон родителя',
        'Паспорт',
        'ИНН',
        'Курсы'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

    students = Abiturient.objects.select_related('school', 'parent').prefetch_related(
        'recordcourse_set__course').filter(is_activ=True)

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 35
    ws.column_dimensions['H'].width = 35
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 40

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, student in enumerate(students, start=1):
        courses = ", ".join([rc.course.name for rc in student.recordcourse_set.all()])

        parent_name = ""
        parent_phone = ""
        parent_pasport = ""
        parent_inn = ""

        if student.parent:
            parent_name = f"{student.parent.last_name} {student.parent.first_name}"
            parent_phone = student.parent.phone or ""
            parent_pasport = student.parent.pasport or ""
            parent_inn = student.parent.inn or ""

        school_name = student.school.name if student.school else ""

        ws.append([
            idx,
            student.last_name or '',
            student.first_name or '',
            student.email or '',
            student.phone or '',
            student.class_name or '',
            school_name,
            parent_name,
            parent_phone,
            parent_pasport,
            parent_inn,
            courses or 'Не записан',
        ])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_students_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)

    return response

# Oтчет по записям на курсы
def generate_enrollment_report(start_date=None, end_date=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Записи на курсы"

    headers = [
        '№',
        'Абитуриент',
        'Email',
        'Телефон',
        'Курс',
        'Преподаватель',
        'Стоимость',
        'Статус',
        'Дата записи'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

    records = RecordCourse.objects.select_related('abiturient', 'course__teacher').order_by('-created_at')

    if start_date and end_date:
        records = records.filter(created_at__range=[start_date, end_date])

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, record in enumerate(records, start=1):
        abiturient_name = f"{record.abiturient.last_name} {record.abiturient.first_name}" if record.abiturient else ""
        abiturient_email = record.abiturient.email if record.abiturient else ""
        abiturient_phone = record.abiturient.phone if record.abiturient else ""

        course_name = record.course.name if record.course else ""
        teacher_name = f"{record.course.teacher.last_name} {record.course.teacher.first_name}" if record.course and record.course.teacher else ""
        course_price = float(record.course.price) if record.course and record.course.price else 0

        # Цвет статуса
        status_display = record.status
        if record.status == 'Подтверждена':
            status_display = "Подтверждена"
        elif record.status == 'Отклонена':
            status_display = "Отклонена"
        elif record.status == 'В обработке':
            status_display = "В обработке"
        elif record.status == 'Завершена':
            status_display = "Завершена"

        ws.append([
            idx,
            abiturient_name,
            abiturient_email,
            abiturient_phone,
            course_name,
            teacher_name,
            course_price,
            status_display,
            record.created_at.strftime('%d.%m.%Y %H:%M') if record.created_at else '',
        ])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_enrollment_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)

    return response

# Oтчет по преподавателям
def generate_teachers_report():
    wb = Workbook()
    ws = wb.active
    ws.title = "Преподаватели"

    headers = [
        '№',
        'Фамилия',
        'Имя',
        'Email',
        'Телефон',
        'Количество курсов',
        'Всего студентов',
        'Статус'
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

    teachers = Teacher.objects.select_related('user').filter(is_activ=True)

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, teacher in enumerate(teachers, start=1):
        courses_count = teacher.course_set.filter(is_activ=True).count()

        total_students = 0
        for course in teacher.course_set.filter(is_activ=True):
            total_students += course.get_enrolled_count()

        status = "Активен" if teacher.is_activ else "Архив"

        ws.append([
            idx,
            teacher.last_name or '',
            teacher.first_name or '',
            teacher.email or '',
            teacher.user.phone if hasattr(teacher.user, 'phone') else '',
            courses_count,
            total_students,
            status,
        ])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=report_teachers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)

    return response